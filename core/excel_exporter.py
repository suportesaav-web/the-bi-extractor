"""Renderizador de arquivo Excel (.xlsx) com estilos corporativos e fórmulas nativas."""

from __future__ import annotations

import io
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import pandas as pd


def generate_excel_workbook(df: pd.DataFrame, sheet_name: str = "Tabela Extraída") -> io.BytesIO:
    """Gera um arquivo Excel (.xlsx) profissional, estilizado e formatado para apresentação executiva.
    
    Suporta tanto a matriz Power BI Saavedra quanto tabelas genéricas extraídas em Modo Híbrido.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Paletas Corporativas
    NAVY_BLUE = "1F4E78"
    WHITE = "FFFFFF"
    ZEBRA_FILL = "F2F4F8"
    TOTAL_FILL = "D9E1F2"
    BORDER_COLOR = "D9D9D9"

    header_font = Font(name="Segoe UI", size=11, bold=True, color=WHITE)
    header_fill = PatternFill(start_color=NAVY_BLUE, end_color=NAVY_BLUE, fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    zebra_pattern = PatternFill(start_color=ZEBRA_FILL, end_color=ZEBRA_FILL, fill_type="solid")
    white_pattern = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
    total_fill_pattern = PatternFill(start_color=TOTAL_FILL, end_color=TOTAL_FILL, fill_type="solid")

    data_font = Font(name="Segoe UI", size=10)
    total_font = Font(name="Segoe UI", size=11, bold=True)

    thin_border_side = Side(border_style="thin", color=BORDER_COLOR)
    data_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    total_top_side = Side(border_style="thin", color=NAVY_BLUE)
    total_bottom_side = Side(border_style="double", color=NAVY_BLUE)
    total_border = Border(top=total_top_side, bottom=total_bottom_side, left=thin_border_side, right=thin_border_side)

    CURRENCY_FORMAT = 'R$ #,##0.00;[Red](R$ #,##0.00);"-"'
    NUMBER_FORMAT = '#,##0.00'
    PERCENT_FORMAT = "0.00%"

    is_powerbi = ("Customer Group" in df.columns and "FY26 PPs" in df.columns)

    # 1. Escrever Cabeçalho
    columns = [str(c) for c in df.columns]
    ws.append(columns)
    ws.row_dimensions[1].height = 28

    for col_num in range(1, len(columns) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = data_border

    # 2. Escrever Dados
    start_row = 2
    for row_idx, row_data in enumerate(df.values, start=start_row):
        ws.append(list(row_data))
        ws.row_dimensions[row_idx].height = 20
        is_even = (row_idx % 2 == 0)
        current_fill = zebra_pattern if is_even else white_pattern

        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.fill = current_fill
            cell.border = data_border

            # Formatação específica Power BI
            if is_powerbi:
                if col_name in ["Customer Group", "Business Unit"]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col_name == "Portfolio":
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                elif col_name in ["FY26 PPs", "Gross Sales Billed", "Gross Sales Open", "Total Gross", "Ating PPs"]:
                    cell.number_format = CURRENCY_FORMAT
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif col_name == "% PPs":
                    cell.number_format = PERCENT_FORMAT
                    cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                # Formatação genérica baseada no tipo do valor
                val = cell.value
                if isinstance(val, (int, float)):
                    cell.number_format = NUMBER_FORMAT
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

    num_data_rows = len(df)
    last_data_row = start_row + num_data_rows - 1
    total_row_idx = last_data_row + 1

    # 3. Linha de Totalização
    if num_data_rows > 0:
        ws.row_dimensions[total_row_idx].height = 24

        if is_powerbi:
            ws.cell(row=total_row_idx, column=1, value="TOTAL GERAL")
            ws.cell(row=total_row_idx, column=2, value="-")
            ws.cell(row=total_row_idx, column=3, value="-")

            for col_idx, col_name in enumerate(columns, start=1):
                col_letter = get_column_letter(col_idx)
                cell = ws.cell(row=total_row_idx, column=col_idx)
                cell.font = total_font
                cell.fill = total_fill_pattern
                cell.border = total_border

                if col_name in ["Customer Group", "Business Unit", "Portfolio"]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col_name in ["FY26 PPs", "Gross Sales Billed", "Gross Sales Open", "Total Gross", "Ating PPs"]:
                    cell.value = f"=SUM({col_letter}{start_row}:{col_letter}{last_data_row})"
                    cell.number_format = CURRENCY_FORMAT
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif col_name == "% PPs":
                    fy26_col_idx = columns.index("FY26 PPs") + 1 if "FY26 PPs" in columns else 4
                    total_gross_col_idx = columns.index("Total Gross") + 1 if "Total Gross" in columns else 7
                    fy26_letter = get_column_letter(fy26_col_idx)
                    gross_letter = get_column_letter(total_gross_col_idx)
                    cell.value = f'=IF({fy26_letter}{total_row_idx}>0, {gross_letter}{total_row_idx}/{fy26_letter}{total_row_idx}, 0)'
                    cell.number_format = PERCENT_FORMAT
                    cell.alignment = Alignment(horizontal="right", vertical="center")
        else:
            # Total para colunas numéricas genéricas
            ws.cell(row=total_row_idx, column=1, value="TOTAL")
            for col_idx, col_name in enumerate(columns, start=1):
                col_letter = get_column_letter(col_idx)
                cell = ws.cell(row=total_row_idx, column=col_idx)
                cell.font = total_font
                cell.fill = total_fill_pattern
                cell.border = total_border

                is_num = pd.api.types.is_numeric_dtype(df[col_name]) if col_name in df.columns else False
                if is_num and col_idx > 1:
                    cell.value = f"=SUM({col_letter}{start_row}:{col_letter}{last_data_row})"
                    cell.number_format = NUMBER_FORMAT
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif col_idx == 1:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.value = "-"
                    cell.alignment = Alignment(horizontal="center", vertical="center")

    # 4. Ajuste Automático da Largura das Colunas
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = cell.value
            if val is not None:
                val_str = str(val)
                if val_str.startswith("="):
                    val_str = "R$ 999.999.999,00"
                max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

    ws.views.sheetView[0].showGridLines = True

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
